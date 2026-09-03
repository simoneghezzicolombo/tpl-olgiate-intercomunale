#!/usr/bin/env python3
"""Build a clearly separated contextual series around the audited 2011-2021 OD trend.

This script does NOT treat the 2019-2023 'popolazione insistente' products as
interchangeable OD matrices. It inventories what the public attachments expose,
extracts the official municipal 2019 resident-side commuting indicators, and
adds recent resident/employment context only where an official municipal SDMX
series is identifiable without assumptions.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs/phase2"
OUT.mkdir(parents=True, exist_ok=True)

MERATESE = {
    "097002": "Airuno",
    "097010": "Brivio",
    "097012": "Calco",
    "097020": "Cernusco Lombardone",
    "097039": "Imbersago",
    "097092": "La Valletta Brianza",
    "097044": "Lomagna",
    "097048": "Merate",
    "097053": "Montevecchia",
    "097058": "Olgiate Molgora",
    "097061": "Osnago",
    "097062": "Paderno d'Adda",
    "097071": "Robbiate",
    "097074": "Santa Maria Hoè",
    "097091": "Verderio",
}

WORKBOOKS = {
    2019: "https://www.istat.it/wp-content/uploads/2022/09/Allegato-statistico-2019_22_07_2022-new.xlsx",
    2020: "https://www.istat.it/wp-content/uploads/2024/11/Allegato-statistico-2020_23_07_2024.xlsx",
    2021: "https://www.istat.it/wp-content/uploads/2026/06/Allegato-statistico-2021_20_03_2026.xlsx",
    2022: "https://www.istat.it/wp-content/uploads/2026/06/Allegato-statistico-2022_20_03_2026.xlsx",
    2023: "https://www.istat.it/wp-content/uploads/2026/06/Allegato-statistico-2023_20_03_2026.xlsx",
}

SDMX_BASE = "https://esploradati.istat.it/SDMXWS/rest/data"
SDMX_ACCEPT = "application/vnd.sdmx.data+csv;version=1.0.0"


def get(url: str, accept: str | None = None) -> bytes:
    headers = {"User-Agent": "tpl-olgiate-phase2-context/1.0"}
    if accept:
        headers["Accept"] = accept
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=300) as response:
        data = response.read()
    if not data:
        raise RuntimeError(f"Empty response: {url}")
    return data


def as_number(value: Any) -> float | None:
    if value is None:
        return None
    s = str(value).strip().replace(" ", "")
    if not s or s in {"-", ".", ".."}:
        return None
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return None


def scan_workbooks() -> dict:
    """Scan every public workbook cell for target codes/names and extract national Tav.1."""
    target_tokens = set(MERATESE)
    for name in MERATESE.values():
        target_tokens.add(name.casefold())
        target_tokens.add(name.replace("'", "’").casefold())
    matches: list[dict] = []
    national: list[dict] = []
    inventory: dict[str, Any] = {}

    for year, url in WORKBOOKS.items():
        data = get(url)
        if data[:2] != b"PK":
            raise RuntimeError(f"Workbook {year} is not XLSX")
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        inventory[str(year)] = {
            "url": url,
            "sha256": hashlib.sha256(data).hexdigest(),
            "sheets": [],
        }
        for ws in wb.worksheets:
            inventory[str(year)]["sheets"].append(
                {"title": ws.title, "rows": ws.max_row, "columns": ws.max_column}
            )
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                vals = list(row)
                # Scan all cells for exact code/name occurrence, but store only matching rows.
                row_strings = ["" if v is None else str(v).strip() for v in vals]
                row_fold = [x.casefold() for x in row_strings]
                hit_tokens: list[str] = []
                for token in target_tokens:
                    if token in row_strings or token in row_fold:
                        hit_tokens.append(token)
                if hit_tokens:
                    matches.append({
                        "year": year,
                        "sheet": ws.title,
                        "row": r_idx,
                        "hits": sorted(set(hit_tokens)),
                        "values": row_strings,
                    })

                # Tav.1 has nationally comparable high-level signal rows.
                if ws.title.startswith("Tav. 1") and row_strings:
                    label = row_strings[0].strip().casefold()
                    if label in {"lavoro", "scuola", "università", "universita"}:
                        # Across editions columns B/C are inside/other municipality.
                        inside = as_number(vals[1] if len(vals) > 1 else None)
                        outside = as_number(vals[2] if len(vals) > 2 else None)
                        national.append({
                            "year": year,
                            "signal": "Università" if label.startswith("univers") else row_strings[0],
                            "inside_own_municipality": inside,
                            "outside_own_municipality": outside,
                            "inside_plus_outside": None if inside is None or outside is None else inside + outside,
                        })
        wb.close()

    (OUT / "mobility_context_workbook_scan.json").write_text(
        json.dumps({"inventory": inventory, "matches": matches}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    with (OUT / "mobility_context_national_2019_2023.csv").open("w", newline="", encoding="utf-8") as f:
        fields = ["year", "signal", "inside_own_municipality", "outside_own_municipality", "inside_plus_outside"]
        w = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        w.writeheader(); w.writerows(national)
    return {"matches": matches, "national": national, "inventory": inventory}


def fetch_sdmx(flow: str, key: str, start: int, end: int) -> tuple[str, list[dict[str, str]]]:
    url = f"{SDMX_BASE}/IT1,{flow},1.0/{key}?startPeriod={start}&endPeriod={end}"
    raw = get(url, SDMX_ACCEPT).decode("utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(raw)))
    if not rows:
        raise RuntimeError(f"No SDMX rows: {flow} {start}-{end}")
    return url, rows


def fetch_2019_municipal_mobility() -> dict:
    codes = "+".join(MERATESE)
    # Dimension order documented by the official SDMX dataflow / AgID ETL:
    # FREQ.REF_AREA.INDICATOR.GENDER.AGE_NOCLASS.CITIZENSHIP.EDU_ATTAIN.CUR_ACT_STAT.LOC_DEST.REAS_COMMUTING
    key = f"A.{codes}.RP_COM_DAY.T.TOTAL.TOTAL.ALL.99.."
    url, rows = fetch_sdmx("DF_DCSS_ISTR_LAV_PEN_2_TV_5", key, 2019, 2019)
    vals: dict[str, dict[tuple[str, str], float]] = {c: {} for c in MERATESE}
    for r in rows:
        c = r.get("REF_AREA", "")
        if c not in vals:
            continue
        try:
            v = float(r["OBS_VALUE"])
        except (KeyError, TypeError, ValueError):
            continue
        vals[c][(r.get("LOC_DEST", ""), r.get("REAS_COMMUTING", ""))] = v

    metrics = [
        ("commuters_total", "ALL", "ALL"),
        ("commuters_outside", "OMPUR", "ALL"),
        ("commuters_inside", "SMPUR", "ALL"),
        ("work_total", "ALL", "WK"),
        ("work_outside", "OMPUR", "WK"),
        ("work_inside", "SMPUR", "WK"),
        ("study_total", "ALL", "STD"),
        ("study_outside", "OMPUR", "STD"),
        ("study_inside", "SMPUR", "STD"),
    ]
    result=[]
    for c,name in MERATESE.items():
        d={"istat":c,"comune":name}
        for label,loc,reason in metrics:
            d[label]=vals[c].get((loc,reason))
        # Algebraic validation is mandatory.
        if d["commuters_total"] != d["commuters_outside"] + d["commuters_inside"]:
            raise RuntimeError(f"2019 inside/outside reconciliation failed: {c}")
        if d["commuters_total"] != d["work_total"] + d["study_total"]:
            raise RuntimeError(f"2019 work/study reconciliation failed: {c}")
        result.append(d)
    with (OUT / "mobility_context_meratese_2019.csv").open("w", newline="", encoding="utf-8") as f:
        w=csv.DictWriter(f,fieldnames=list(result[0]),lineterminator="\n");w.writeheader();w.writerows(result)
    return {"url":url,"rows":result}


def fetch_recent_population_and_employment() -> dict:
    """Fetch recent municipal residents and inspect employment dataflow without guessing categories."""
    codes = "+".join(MERATESE)
    pop_url, pop_rows = fetch_sdmx("DF_DCSS_FAM_POP_TV_1", f"A.{codes}.RESPOP_AV", 2019, 2024)
    pop: dict[tuple[str,int],float]={}
    for r in pop_rows:
        try:
            c=r["REF_AREA"]; y=int(str(r["TIME_PERIOD"])[:4]); v=float(r["OBS_VALUE"])
        except (KeyError,ValueError,TypeError):
            continue
        if c in MERATESE: pop[(c,y)]=v

    # Pull labour-condition data for 2019-2024. We only materialise rows whose
    # category semantics are already documented in the repo/official flow:
    # CUR_ACT_STAT=1 occupied, GENDER=T, CITIZENSHIP=TOTAL, EDU=ALL, LOC=ALL,
    # REAS=ALL. The published aggregate Y_GE15 means age 15 and over and is
    # directly usable as the total employed-resident population. We do not sum
    # the overlapping age categories.
    lab_url, lab_rows = fetch_sdmx("DF_DCSS_ISTR_LAV_PEN_2_TV_3", f"A.{codes}........", 2019, 2024)
    occupied: dict[tuple[str,int],float]={}
    age_values=set()
    for r in lab_rows:
        if r.get("REF_AREA") not in MERATESE or r.get("CUR_ACT_STAT") != "1":
            continue
        if r.get("GENDER") != "T" or r.get("CITIZENSHIP") != "TOTAL" or r.get("EDU_ATTAIN") != "ALL" or r.get("LOC_DEST") != "ALL" or r.get("REAS_COMMUTING") != "ALL":
            continue
        age_values.add(r.get("AGE_NOCLASS",""))
        if r.get("AGE_NOCLASS") != "Y_GE15":
            continue
        try:
            occupied[(r["REF_AREA"],int(str(r["TIME_PERIOD"])[:4]))]=float(r["OBS_VALUE"])
        except (KeyError,ValueError,TypeError): pass

    years=sorted({y for _,y in pop}|{y for _,y in occupied})
    out=[]
    for c,name in MERATESE.items():
        for y in years:
            if (c,y) not in pop and (c,y) not in occupied: continue
            out.append({"istat":c,"comune":name,"year":y,"resident_population":pop.get((c,y)),"employed_residents":occupied.get((c,y))})
    with (OUT / "mobility_context_population_employment_2019_2024.csv").open("w",newline="",encoding="utf-8") as f:
        fields=["istat","comune","year","resident_population","employed_residents"]
        w=csv.DictWriter(f,fieldnames=fields,lineterminator="\n");w.writeheader();w.writerows(out)
    return {"population_url":pop_url,"labour_url":lab_url,"available_age_codes_for_occupied":sorted(age_values),"occupied_total_identified":bool(occupied),"rows":out}


def main() -> int:
    scan=scan_workbooks()
    m2019=fetch_2019_municipal_mobility()
    recent=fetch_recent_population_and_employment()
    report={
        "scope_note":"Context only. These outputs are not merged into the audited 2011-2021 OD series unless the underlying universe is demonstrated compatible.",
        "public_workbook_target_match_count":len(scan["matches"]),
        "municipal_2019_sdmx":"IDENTIFIED_AND_VALIDATED",
        "recent_population_employment":{
            "occupied_total_identified":recent["occupied_total_identified"],
            "available_age_codes_for_occupied":recent["available_age_codes_for_occupied"],
        },
        "student_recent_municipal_series":"NOT_IDENTIFIED_IN_PUBLIC_EQUIVALENT_SERIES",
        "sources":{
            "workbooks":WORKBOOKS,
            "municipal_2019":m2019["url"],
            "population":recent["population_url"],
            "labour":recent["labour_url"],
        },
        "outputs":[
            "outputs/phase2/mobility_context_workbook_scan.json",
            "outputs/phase2/mobility_context_national_2019_2023.csv",
            "outputs/phase2/mobility_context_meratese_2019.csv",
            "outputs/phase2/mobility_context_population_employment_2019_2024.csv",
        ],
    }
    (OUT/"mobility_context_audit.json").write_text(json.dumps(report,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    print(json.dumps(report,ensure_ascii=False,indent=2))
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
