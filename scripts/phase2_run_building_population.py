#!/usr/bin/env python3
"""Canonical fail-closed entrypoint for Phase 2 building population.

The current official ISTAT 2023 regional package is identified from the package
contents, not from a guessed filename. Building allocations are spatially split
by census section and accessibility is evaluated at a representative point of
each building-section intersection, preserving boundary integrity.
"""
from __future__ import annotations

import hashlib
import io
import math
from pathlib import Path
import sys
import zipfile

import geopandas as gpd
from openpyxl import load_workbook
import pandas as pd
from shapely.geometry import Point

import phase2_build_building_population as impl


SECTION_EXACT = (
    "SEZ21_ID", "SEZ2021_ID", "SEZIONE", "COD_SEZ", "SEZ23_ID", "SEZ_ID"
)
MUNICIPALITY_EXACT = (
    "PRO_COM", "PRO_COM_T", "COD_COM", "CODICE_COMUNE", "COD_COMUNE"
)
POPULATION_EXACT = (
    "P1", "POP2023", "POP23", "POP_TOT", "POP", "POPOLAZIONE"
)


def _single(items: list, label: str):
    if len(items) != 1:
        raise RuntimeError(f"expected exactly one {label}, got {items}")
    return items[0]


def _workbook_member(names: list[str]) -> str:
    preferred = [
        n for n in names
        if n.lower().endswith(".xlsx")
        and "r03" in n.lower()
        and "tracciato" not in n.lower()
        and ("sez" in n.lower() or "indicator" in n.lower())
    ]
    if len(preferred) == 1:
        return preferred[0]
    fallback = [
        n for n in names
        if n.lower().endswith(".xlsx")
        and "r03" in n.lower()
        and "tracciato" not in n.lower()
    ]
    return _single(fallback, "Lombardia regional section workbook in official 2023 package")


def _find_exact(columns, names):
    lookup = {str(c).strip().upper(): c for c in columns if c is not None}
    for name in names:
        if name in lookup:
            return lookup[name]
    return None


def _section_candidates(columns) -> list:
    exact = _find_exact(columns, SECTION_EXACT)
    if exact is not None:
        return [exact]
    out = []
    for c in columns:
        low = str(c).strip().lower()
        if "sez" in low and any(k in low for k in ("id", "cod", "2021", "21")):
            out.append(c)
    return out


def _municipality_candidates(columns) -> list:
    exact = _find_exact(columns, MUNICIPALITY_EXACT)
    if exact is not None:
        return [exact]
    out = []
    for c in columns:
        low = str(c).strip().lower()
        if "pro_com" in low or ("cod" in low and "com" in low):
            out.append(c)
    return out


def _population_candidates(columns) -> list:
    exact = _find_exact(columns, POPULATION_EXACT)
    if exact is not None:
        return [exact]
    out = []
    for c in columns:
        low = str(c).strip().lower()
        if "popol" in low and ("resident" in low or low == "popolazione"):
            out.append(c)
    return out


def _discover_table(raw: bytes) -> tuple[pd.DataFrame, str, int, object, object, object]:
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    matches = []
    observations = []
    try:
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for header in range(0, 11):
                values = next(
                    ws.iter_rows(min_row=header + 1, max_row=header + 1, values_only=True),
                    None,
                )
                if not values:
                    continue
                columns = list(values)
                sec = _section_candidates(columns)
                muni = _municipality_candidates(columns)
                pop = _population_candidates(columns)
                if sec or muni or pop:
                    observations.append({
                        "sheet": sheet,
                        "header": header,
                        "section": [str(v) for v in sec],
                        "municipality": [str(v) for v in muni],
                        "population": [str(v) for v in pop],
                    })
                if len(sec) == 1 and len(pop) == 1 and len(muni) == 1:
                    matches.append((sheet, header, sec[0], muni[0], pop[0]))
                    break
    finally:
        workbook.close()

    if len(matches) != 1:
        raise RuntimeError(
            "ISTAT 2023 workbook schema is not uniquely identifiable with an explicit municipal key; "
            f"matches={[(m[0], m[1], str(m[2]), str(m[3]), str(m[4])) for m in matches]}; "
            f"candidate_observations={observations[:30]}"
        )
    sheet, header, section_col, muni_col, pop_col = matches[0]
    df = pd.read_excel(io.BytesIO(raw), sheet_name=sheet, header=header, dtype=object, engine="openpyxl")
    return df, sheet, header, section_col, muni_col, pop_col


def load_istat_2023_sections(source_dir: Path, selected_codes: set[str]):
    zip_path = source_dir / "istat_sections_2023.zip"
    info = impl.download(impl.ISTAT_SECTIONS_2023_URL, zip_path)
    with zipfile.ZipFile(zip_path) as z:
        names = z.namelist()
        member = _workbook_member(names)
        raw = z.read(member)

    df, sheet, header, section_col, muni_col, pop_col = _discover_table(raw)
    xlsx_path = source_dir / "R03_2023_sections_selected_source.xlsx"
    xlsx_path.write_bytes(raw)

    section_raw = df[section_col].astype(str)
    out = pd.DataFrame({
        "section_id_raw": section_raw,
        "section_id": df[section_col].map(impl.normalise_section),
        "municipality_code": df[muni_col].map(impl.normalise_municipality),
        "population_2023_fact": pd.to_numeric(df[pop_col], errors="coerce"),
    })
    out = out.loc[out["municipality_code"].isin(selected_codes)].copy()
    if out.empty:
        raise RuntimeError("no selected municipalities found in ISTAT 2023 section table")
    if out["population_2023_fact"].isna().any():
        bad = out.loc[out["population_2023_fact"].isna(), ["section_id_raw", "municipality_code"]].head(20)
        raise RuntimeError(f"non-numeric ISTAT 2023 population in selected rows: {bad.to_dict('records')}")
    if (out["population_2023_fact"] < 0).any():
        raise RuntimeError("negative ISTAT 2023 section population")
    if not out["population_2023_fact"].map(math.isfinite).all():
        raise RuntimeError("non-finite ISTAT 2023 section population")
    if out["section_id"].eq("").any():
        raise RuntimeError("blank section identifier in selected ISTAT 2023 rows")
    if out["section_id"].duplicated().any():
        dup = out.loc[out["section_id"].duplicated(keep=False), "section_id"].head(20).tolist()
        raise RuntimeError(f"duplicate ISTAT 2023 section IDs: {dup}")
    missing_muni = selected_codes - set(out["municipality_code"])
    if missing_muni:
        raise RuntimeError(f"selected municipalities absent from ISTAT 2023: {sorted(missing_muni)}")

    out["is_fictitious_section"] = out["section_id_raw"].map(impl.is_fictitious_section)
    out["population_2023_epistemic_status"] = "FACT_ISTAT_CENSUS_SECTION_2023"
    info.update({
        "epistemic_status": "FACT",
        "reference_year": 2023,
        "zip_member": member,
        "workbook_sheet": sheet,
        "header_row_zero_based": header,
        "regional_xlsx_sha256": hashlib.sha256(raw).hexdigest(),
        "regional_xlsx_bytes": len(raw),
        "section_id_field": str(section_col),
        "municipality_field_or_method": f"FACT_FIELD:{muni_col}",
        "population_field": str(pop_col),
        "selected_rows": len(out),
        "selected_fictitious_rows": int(out["is_fictitious_section"].sum()),
    })
    return out, info


def build_section_pieces(buildings: gpd.GeoDataFrame, section_geometry: gpd.GeoDataFrame) -> pd.DataFrame:
    eligible = buildings.loc[
        buildings["eligible_primary"] | buildings["eligible_fallback"],
        [
            "CLASSREF", "footprint_area_m2", "dbgt_volume_complete", "dbgt_volume_proxy_m3",
            "eligible_primary", "eligible_fallback", "allocation_weight_basis", "geometry",
        ],
    ].copy()
    sec = section_geometry[["section_id", "municipality_code", "geometry"]].copy()
    joined = gpd.sjoin(eligible, sec, how="inner", predicate="intersects")
    sec_geom = sec.geometry.to_dict()
    rows = []
    for row in joined.itertuples():
        inter = row.geometry.intersection(sec_geom[row.index_right])
        area = float(inter.area) if not inter.is_empty else 0.0
        if area <= 0:
            continue
        piece_point = inter.representative_point()
        footprint_area = float(row.footprint_area_m2)
        if bool(row.dbgt_volume_complete) and pd.notna(row.dbgt_volume_proxy_m3) and footprint_area > 0:
            weight = float(row.dbgt_volume_proxy_m3) * area / footprint_area
            basis = "DBGT_VOLUME_PROXY_COMPLETE_PRORATED_BY_SECTION_INTERSECTION"
        else:
            weight = area
            basis = "DBGT_FOOTPRINT_SECTION_INTERSECTION_AREA"
        rows.append({
            "building_id": row.CLASSREF,
            "section_id": row.section_id,
            "municipality_code": row.municipality_code,
            "eligible_primary": bool(row.eligible_primary),
            "eligible_fallback": bool(row.eligible_fallback),
            "intersection_area_m2": area,
            "allocation_weight": weight,
            "allocation_weight_basis_piece": basis,
            "piece_x_utm32": float(piece_point.x),
            "piece_y_utm32": float(piece_point.y),
            "piece_representative_point_epistemic_status": "DERIVED_FROM_DBGT_SECTION_INTERSECTION",
            "weight_epistemic_status": "DERIVED_FROM_DBGT_GEOMETRY_AND_AVAILABLE_VOLUME",
        })
    return pd.DataFrame(rows)


def compute_accessibility(
    building_allocations: pd.DataFrame,
    building_points: pd.DataFrame,
    gate_b_dir: Path,
) -> pd.DataFrame:
    del building_points
    _, _, _, network_minutes, tree, node_ids, _ = impl.build_gate_b_access(gate_b_dir)
    core_alloc = building_allocations.loc[
        building_allocations["municipality_code"].isin(impl.CORE_CODES)
    ].copy()
    required = {"piece_x_utm32", "piece_y_utm32"}
    if not required.issubset(core_alloc.columns):
        raise RuntimeError(f"building-section piece coordinates missing: {required - set(core_alloc.columns)}")
    distances, indexes = tree.query(core_alloc[["piece_x_utm32", "piece_y_utm32"]].to_numpy(), k=1)
    core_alloc["nearest_graph_node_id"] = [node_ids[int(i)] for i in indexes]
    core_alloc["connector_distance_m"] = [float(v) for v in distances]
    core_alloc["connector_walk_min"] = core_alloc["connector_distance_m"] / impl.CONNECTOR_M_PER_MIN
    core_alloc["connector_within_limit"] = core_alloc["connector_distance_m"] <= impl.POP_CONNECTOR_MAX_M
    core_alloc["network_walk_min_to_gtfs_stop"] = core_alloc["nearest_graph_node_id"].map(network_minutes)
    core_alloc["walk_min_to_nearest_gtfs_stop"] = (
        core_alloc["network_walk_min_to_gtfs_stop"] + core_alloc["connector_walk_min"]
    )
    core_alloc.loc[
        ~core_alloc["connector_within_limit"] | core_alloc["network_walk_min_to_gtfs_stop"].isna(),
        "walk_min_to_nearest_gtfs_stop",
    ] = math.nan
    for threshold in (5, 8, 10, 12):
        core_alloc[f"covered_{threshold}min"] = core_alloc[
            "walk_min_to_nearest_gtfs_stop"
        ].le(threshold).fillna(False)
    core_alloc["accessibility_epistemic_status"] = (
        "MODEL_OUTPUT_GATE_B_WALK_GRAPH_BUILDING_SECTION_PIECE_REPRESENTATIVE_POINT"
    )
    return core_alloc


def boundary_comparison(
    building_allocations: pd.DataFrame,
    building_points: pd.DataFrame,
    gate_b_dir: Path,
    core_boundaries_path: Path,
) -> pd.DataFrame:
    del building_points
    boundaries = gpd.read_file(core_boundaries_path).to_crs(32632)
    boundaries["municipality_code"] = boundaries["PRO_COM_T"].map(impl.normalise_municipality)
    boundary_map = boundaries.set_index("municipality_code").geometry.to_dict()

    cells = pd.read_csv(gate_b_dir / "population_cells_real.csv")
    cells["municipality_code"] = cells["PRO_COM_T"].map(impl.normalise_municipality)
    cell_pts = gpd.GeoDataFrame(cells, geometry=gpd.points_from_xy(cells.lon, cells.lat), crs=4326).to_crs(32632)
    cell_pts["distance_to_municipal_boundary_m"] = [
        row.geometry.distance(boundary_map[row.municipality_code].boundary)
        for row in cell_pts.itertuples()
    ]
    v1_band = cell_pts.loc[
        cell_pts["distance_to_municipal_boundary_m"] <= impl.BOUNDARY_COMPARISON_BAND_M
    ]
    v1 = v1_band.groupby("municipality_code")["pop_calibrated_2025"].sum().to_dict()

    alloc = building_allocations.loc[
        building_allocations["municipality_code"].isin(impl.CORE_CODES)
    ].copy()
    alloc["distance_to_municipal_boundary_m"] = [
        Point(float(r.piece_x_utm32), float(r.piece_y_utm32)).distance(
            boundary_map[r.municipality_code].boundary
        )
        for r in alloc.itertuples()
    ]
    v2_band = alloc.loc[
        alloc["distance_to_municipal_boundary_m"] <= impl.BOUNDARY_COMPARISON_BAND_M
    ]
    v2 = v2_band.groupby("municipality_code")["building_piece_population_model"].sum().to_dict()
    return pd.DataFrame([
        {
            "municipality_code": code,
            "boundary_band_m_assumption": impl.BOUNDARY_COMPARISON_BAND_M,
            "v1_worldpop_population_near_boundary": float(v1.get(code, 0.0)),
            "v2_building_population_near_boundary": float(v2.get(code, 0.0)),
            "v2_minus_v1_population_near_boundary": float(v2.get(code, 0.0) - v1.get(code, 0.0)),
            "epistemic_status": "MODEL_OUTPUT_SPATIAL_COMPARISON",
        }
        for code in sorted(impl.CORE_CODES)
    ])


def spatial_distribution_comparison(
    building_allocations: pd.DataFrame,
    building_points: pd.DataFrame,
    gate_b_dir: Path,
) -> pd.DataFrame:
    del building_points
    cells = pd.read_csv(gate_b_dir / "population_cells_real.csv")
    cells["municipality_code"] = cells["PRO_COM_T"].map(impl.normalise_municipality)
    cells_gdf = gpd.GeoDataFrame(
        cells,
        geometry=gpd.points_from_xy(cells.lon, cells.lat),
        crs=4326,
    ).to_crs(32632)
    cells_gdf["x"] = cells_gdf.geometry.x
    cells_gdf["y"] = cells_gdf.geometry.y
    alloc = building_allocations.loc[
        building_allocations["municipality_code"].isin(impl.CORE_CODES)
    ].copy()
    rows = []
    for code in sorted(impl.CORE_CODES) + ["CORE_TOTAL"]:
        v1 = cells_gdf if code == "CORE_TOTAL" else cells_gdf.loc[cells_gdf["municipality_code"] == code]
        v2 = alloc if code == "CORE_TOTAL" else alloc.loc[alloc["municipality_code"] == code]
        w1 = v1["pop_calibrated_2025"].astype(float)
        w2 = v2["building_piece_population_model"].astype(float)
        x1 = float((v1["x"] * w1).sum() / w1.sum())
        y1 = float((v1["y"] * w1).sum() / w1.sum())
        x2 = float((v2["piece_x_utm32"] * w2).sum() / w2.sum()) if w2.sum() else math.nan
        y2 = float((v2["piece_y_utm32"] * w2).sum() / w2.sum()) if w2.sum() else math.nan
        shift = math.hypot(x2 - x1, y2 - y1) if math.isfinite(x2) else math.nan
        rows.append({
            "municipality_code": code,
            "v1_worldpop_weighted_centroid_x_utm32": x1,
            "v1_worldpop_weighted_centroid_y_utm32": y1,
            "v2_building_weighted_centroid_x_utm32": x2,
            "v2_building_weighted_centroid_y_utm32": y2,
            "weighted_centroid_shift_m": shift,
            "epistemic_status": "DERIVED_SPATIAL_DISTRIBUTION_COMPARISON",
        })
    return pd.DataFrame(rows)


impl.load_istat_2023_sections = load_istat_2023_sections
impl.build_section_pieces = build_section_pieces
impl.compute_accessibility = compute_accessibility
impl.boundary_comparison = boundary_comparison
impl.spatial_distribution_comparison = spatial_distribution_comparison

if __name__ == "__main__":
    sys.exit(impl.main())
