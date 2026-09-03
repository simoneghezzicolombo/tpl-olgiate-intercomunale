#!/usr/bin/env python3
"""Read-only fail-closed schema probe for official building-population sources."""
from __future__ import annotations

import io
import json
import zipfile
from pathlib import Path

import geopandas as gpd
from openpyxl import load_workbook
import pandas as pd
import requests

ISTAT_GEOM_URL = "https://www.istat.it/storage/cartografia/basi_territoriali/2021/R03_21.zip"
ISTAT_DATA_2023_URL = "https://esploradati.istat.it/databrowser/DWL/PERMPOP/SUBCOM/Dati_regionali_2023.zip"
DBGT_BASE = "https://www.cartografia.servizirl.it/arcgis5/rest/services/BaseMap/DBGT_Tema0201_Edificato/MapServer"
CORE_CODES = {"097010", "097012", "097058", "097074", "097092"}
HEADERS = {
    "User-Agent": (
        "tpl-olgiate-building-population/1.0 "
        "(+https://github.com/simoneghezzicolombo/tpl-olgiate-intercomunale)"
    )
}


def get(url: str, *, params: dict | None = None, timeout: int = 240) -> requests.Response:
    r = requests.get(url, params=params, headers=HEADERS, timeout=timeout)
    r.raise_for_status()
    return r


def _normalise_code(series: pd.Series, width: int = 6) -> pd.Series:
    return series.astype(str).str.replace(r"\.0$", "", regex=True).str.strip().str.zfill(width)


def _select_lombardia_workbook(names: list[str]) -> str:
    # Release year is established by the official package URL. Internal filenames
    # need not repeat it.
    candidates = [
        n for n in names
        if n.lower().endswith(".xlsx")
        and "r03" in n.lower()
        and "tracciato" not in n.lower()
        and ("sez" in n.lower() or "indicator" in n.lower())
    ]
    if len(candidates) != 1:
        raise RuntimeError(
            "Could not identify exactly one Lombardia section workbook in official 2023 package; "
            f"candidates={candidates}; xlsx={[n for n in names if n.lower().endswith('.xlsx')]}"
        )
    return candidates[0]


def _candidate_fields(columns: list[object]) -> dict:
    names = [str(c).strip() for c in columns if c is not None]
    sections = [c for c in names if "sez" in c.lower()]
    municipalities = [
        c for c in names
        if "pro_com" in c.lower() or ("cod" in c.lower() and "com" in c.lower())
    ]
    population = [
        c for c in names
        if c.upper() in {"P1", "POP", "POP2023", "POP23", "POP_TOT", "POPOLAZIONE"}
        or ("popol" in c.lower() and "resident" in c.lower())
    ]
    return {"section": sections, "municipality": municipalities, "population": population}


def inspect_istat() -> dict:
    geom = get(ISTAT_GEOM_URL).content
    data = get(ISTAT_DATA_2023_URL).content
    out: dict = {"geometry_bytes": len(geom), "data_2023_bytes": len(data)}

    with zipfile.ZipFile(io.BytesIO(geom)) as z:
        out["geometry_members"] = z.namelist()
        z.extractall("/tmp/istat_geom")
    shp = next(Path("/tmp/istat_geom").rglob("*.shp"))
    gdf = gpd.read_file(shp)
    out["geometry_columns"] = list(gdf.columns)
    out["geometry_crs"] = str(gdf.crs)
    out["geometry_rows"] = int(len(gdf))
    code_col = next((c for c in ("PRO_COM_T", "PRO_COM") if c in gdf.columns), None)
    if code_col is None:
        raise RuntimeError(f"ISTAT geometry lacks municipal code: {list(gdf.columns)}")
    codes = _normalise_code(gdf[code_col])
    out["geometry_rows_province_097"] = int(codes.str.startswith("097").sum())
    out["geometry_rows_core"] = int(codes.isin(CORE_CODES).sum())
    core = gdf.loc[codes.isin(CORE_CODES)].copy().to_crs(7791)
    minx, miny, maxx, maxy = core.total_bounds
    out["core_bbox_epsg7791"] = [float(minx), float(miny), float(maxx), float(maxy)]

    with zipfile.ZipFile(io.BytesIO(data)) as z:
        names = z.namelist()
        out["data_2023_members"] = names
        member = _select_lombardia_workbook(names)
        raw = z.read(member)
    out["data_2023_selected_member"] = member

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    observations = []
    selected = []
    try:
        out["data_2023_sheets"] = wb.sheetnames
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for header in range(0, 11):
                row = next(ws.iter_rows(min_row=header + 1, max_row=header + 1, values_only=True), None)
                if not row:
                    continue
                cand = _candidate_fields(list(row))
                if any(cand.values()):
                    obs = {"sheet": sheet, "header": header, **cand}
                    observations.append(obs)
                    if len(cand["section"]) >= 1 and len(cand["population"]) >= 1:
                        selected.append(obs)
                        break
    finally:
        wb.close()
    if len(selected) != 1:
        raise RuntimeError(f"ISTAT 2023 table header not uniquely identified: {selected}")
    choice = selected[0]
    out["data_2023_selected_sheet"] = choice["sheet"]
    out["data_2023_header_row_zero_based"] = choice["header"]
    out["section_key_candidates"] = choice["section"]
    out["municipality_key_candidates"] = choice["municipality"]
    out["population_field_candidates"] = choice["population"]
    out["header_candidate_observations"] = observations
    return out


def arc_query(layer: int, params: dict) -> dict:
    payload = get(f"{DBGT_BASE}/{layer}/query", params={"f": "json", **params}).json()
    if "error" in payload:
        raise RuntimeError(json.dumps(payload["error"], ensure_ascii=False))
    return payload


def inspect_dbgt(core_bbox: list[float]) -> dict:
    out: dict = {}
    service = get(DBGT_BASE, params={"f": "pjson"}).json()
    out["service_spatial_reference"] = service.get("spatialReference")
    for layer in (0, 3, 5, 22, 24):
        meta = get(f"{DBGT_BASE}/{layer}", params={"f": "pjson"}).json()
        out[f"layer_{layer}_name"] = meta.get("name")
        out[f"layer_{layer}_fields"] = [
            {"name": f.get("name"), "alias": f.get("alias"), "domain": f.get("domain")}
            for f in meta.get("fields", [])
        ]
        out[f"layer_{layer}_relationships"] = meta.get("relationships")

    minx, miny, maxx, maxy = core_bbox
    envelope = f"{minx-5000},{miny-5000},{maxx+5000},{maxy+5000}"
    ids = arc_query(3, {
        "where": "1=1",
        "geometry": envelope,
        "geometryType": "esriGeometryEnvelope",
        "inSR": "7791",
        "spatialRel": "esriSpatialRelIntersects",
        "returnIdsOnly": "true",
    }).get("objectIds") or []
    out["footprint_layer3_ids_5km_envelope"] = len(ids)
    sample_ids = ids[:20]
    if not sample_ids:
        return out
    sample = arc_query(3, {
        "objectIds": ",".join(map(str, sample_ids)),
        "outFields": "OBJECTID,CLASSREF,COD_CONS,DATA_FIN",
        "returnGeometry": "false",
    })
    attrs = [f["attributes"] for f in sample.get("features", [])]
    out["footprint_sample"] = attrs
    refs = [str(a["CLASSREF"]) for a in attrs if a.get("CLASSREF")]
    if not refs:
        return out
    safe = ",".join("'" + r.replace("'", "''") + "'" for r in refs)
    for layer, key, label in ((22, "CLASSID", "building_table_sample"), (24, "CLASSREF", "use_table_sample")):
        payload = arc_query(layer, {
            "where": f"{key} IN ({safe})",
            "outFields": "*",
            "returnGeometry": "false",
        })
        out[label] = [f["attributes"] for f in payload.get("features", [])]
    volumes = arc_query(0, {
        "where": f"CEDIUV IN ({safe})",
        "outFields": "CLASSID,CEDIUV,UN_VOL_AV,UN_VOL_EX,UN_VOL_QE,Shape_Area,DATA_FIN",
        "returnGeometry": "false",
    })
    out["volume_unit_sample"] = [f["attributes"] for f in volumes.get("features", [])]
    return out


def main() -> None:
    istat = inspect_istat()
    bbox = istat.get("core_bbox_epsg7791")
    if not bbox:
        raise SystemExit("Could not derive core bbox from official ISTAT geometry")
    dbgt = inspect_dbgt(bbox)
    print(json.dumps({"istat": istat, "dbgt": dbgt}, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
